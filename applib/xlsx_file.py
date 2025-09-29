from openpyxl import Workbook
import openpyxl
from pathlib import Path
from logging import Logger


class XlsxFile:
    COLUMNS = ('id', 'url', 'comments', 'subs',)

    def __init__(self, logger: Logger, filepath: str) -> None:
        self._filepath = filepath
        self._logger = logger
        self._prepare_file()

    

    def add_row(self, data: dict[str, str]) -> None:
        """
        Add row to end of xlsx file. Where keys are column names and values  - row data.
        """
        workbook = openpyxl.load_workbook(self._filepath)
        
        sheet = workbook.active

        if not sheet:
            self._logger.warning('No active sheet to add row')
            return

        sheet.append(list(data.values()))
        
        workbook.save(self._filepath)


    def rows(self) -> tuple:
        """
        Return rows from xlsx file.
        """
        wb = openpyxl.load_workbook(self._filepath)

        # Select the active sheet
        sheet = wb.active
        
        if not sheet:
            self._logger.warning('no active sheet. Return an empty tuple.')
            return tuple()
        
        rows = []

        # Read and print the data
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            rows.append(row)
    
        return tuple(rows)


    def _prepare_file(self) -> None:
        file_path = Path(self._filepath)
        
        file_path.parent.mkdir(parents=True, exist_ok=True)

        if not Path.exists(file_path):
            wb = Workbook()
            sheet = wb.active
            if not sheet:
                self._logger.warning('Could not found active sheet.')
                return
            sheet.append(self.COLUMNS)
            wb.save(self._filepath)
